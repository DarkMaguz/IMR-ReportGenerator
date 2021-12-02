from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table

import sys
import json

import utils
import settings as cfg


def loadURLs(excelFile):
	urlList = []
	try:
		wb = load_workbook(filename=excelFile, read_only=True)
		sheet = wb.active
		for row in sheet.iter_rows(min_col=38, max_col=39, min_row=1, max_row=1):
			if row[0].value != 'Pdf_URL':
				raise 'Invalid Excel file - expected column "AL1" to contain "Pdf_URL"'
			if row[1].value != 'Report Html Address':
				raise 'Invalid Excel file - expected column "AM1" to contain "Report Html Address"'
		for row in sheet.iter_rows(min_col=38, max_col=39, min_row=cfg.minURLs, max_row=cfg.maxURLs, values_only=True):
			urlList.append({'BRN': '', 'URLs': [row[0], row[1]]})
		for index, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, min_row=cfg.minURLs, max_row=cfg.maxURLs, values_only=True)):
			urlList[index]['BRN'] = row[0]
	except Exception as e:
		print('loadURLs() error loading urls:', e)
		sys.exit(1)
	finally:
		wb.close()
	return urlList


def writeJSON(dlReports):
	with open(cfg.jsonFile, 'w') as file:
		json.dump(dlReports, file, indent=2, sort_keys=True)


def writeReport(dlReports):
	writeJSON(dlReports)
	#return
	try:
		wb = Workbook()
		sheet = wb.active
		sheet.title = 'downloads'
		rowIterator = 1
		sheet.cell(row=rowIterator, column=1, value='BRnum')
		sheet.cell(row=rowIterator, column=2, value='Was downloaded')
		sheet.cell(row=rowIterator, column=3, value='Name of document file')
		for report in dlReports:
			rowIterator += 1
			dlSuccess = 'no'
			fileName = ''
			for data in report['dataPair']:
				if data['isOK']:
					dlSuccess = 'yes'
					fileName = data['fileName']
					break
			sheet.cell(row=rowIterator, column=1, value=report['BRN'])
			sheet.cell(row=rowIterator, column=2, value=dlSuccess)
			sheet.cell(row=rowIterator, column=3, value=fileName)
		wb.save(filename=cfg.metadataFile)
	except Exception as e:
		print('writeReport() error writing report:', e)
		sys.exit(1)
